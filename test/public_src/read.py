__author__ = '76561'

from openpyxl import load_workbook
import configparser

class Readlines:
    def __init__(self,test,sheek,sheek_2):
        self.test=test
        self.sheek=sheek
        self.sheek_2=sheek_2
    def read(self,config,section,option,option_2):
        wb=load_workbook(self.test)
        sheet=wb[self.sheek]
        res=self.read_2()
        res_2=res+1#读一次就+1
        mode=eval(ReadConfig().Config(config,section,option))
        list_1=[]
        if mode==0:
            for item in range(2,sheet.max_row+1):
                dict_1={}
                dict_1[sheet.cell(1,2).value]=sheet.cell(item,2).value
                dict_1[sheet.cell(1,3).value]=sheet.cell(item,3).value
                dict_1[sheet.cell(1,4).value]=sheet.cell(item,4).value
                if sheet.cell(item,5).value.find('${Initializatio}')!=-1:
                    new=sheet.cell(item,5).value.replace('${Initializatio}',str(res))
                    dict_1[sheet.cell(1,5).value]=eval(new)
                else:
                    dict_1[sheet.cell(1,5).value]=eval(sheet.cell(item,5).value)
                dict_1[sheet.cell(1,6).value]=sheet.cell(item,6).value
                dict_1[sheet.cell(1,7).value]=sheet.cell(item,7).value
                list_1.append(dict_1)
            self.update(res_2)#读一次就+1
            return list_1
        elif mode==1:
            re=ReadConfig().Config(config,section,option_2)
            for item in eval(re):
                dict_1={}
                dict_1[sheet.cell(1,2).value]=sheet.cell(item+1,2).value
                dict_1[sheet.cell(1,3).value]=sheet.cell(item+1,3).value
                dict_1[sheet.cell(1,4).value]=sheet.cell(item+1,4).value
                if sheet.cell(item+1,5).value.find('${Initializatio}')!=-1:
                    new=sheet.cell(item+1,5).value.replace('${Initializatio}',str(res))
                    dict_1[sheet.cell(1,5).value]=eval(new)
                else:
                    dict_1[sheet.cell(1,5).value]=eval(sheet.cell(item+1,5).value)
                dict_1[sheet.cell(1,6).value]=sheet.cell(item+1,6).value
                dict_1[sheet.cell(1,7).value]=sheet.cell(item+1,7).value
                list_1.append(dict_1)
            self.update(res_2)#读一次就+1的值更新到表单里
            return list_1
    def read_2(self):
        wb=load_workbook(self.test)
        sheet=wb[self.sheek_2]
        re=sheet.cell(1,2).value
        return re
    def update(self,res_l):
        wb=load_workbook(self.test)
        sheet=wb[self.sheek_2]
        sheet.cell(1,2).value=res_l
        wb.save(self.test)
    def Write_test_results(self,x,y,va):
        wb=load_workbook(self.test)
        sheet=wb[self.sheek]
        sheet.cell(x,y).value=va
        wb.save(self.test)
class ReadConfig:
    def Config(self,config,section,option):
        cf=configparser.ConfigParser()#实例
        cf.read(config)
        re=cf.get(section,option)
        return re
if __name__ == '__main__':
    from config import path
    res=Readlines(path.test_date,'register','init_date').read(path.confg_path,'FLAG','mode','case_id')
    print(res)